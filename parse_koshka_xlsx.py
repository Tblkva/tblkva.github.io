import json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit(
        "Не удалось импортировать openpyxl. "
        "Установите его командой: python -m pip install openpyxl"
    ) from e


def cell_repr(cell):
    """Короткое текстовое представление ячейки."""
    value = cell.value
    if value is None:
        return None
    text = str(value)
    # Обрезаем очень длинные значения
    if len(text) > 80:
        text = text[:77] + "..."
    return text


def parse_xlsx(relative_dir: str, filename: str) -> None:
    """
    Распарсить один XLSX-файл (кошка/собака) в .parsed.json.

    relative_dir: подкаталог (например, 'кошка' или 'собака')
    filename: имя файла XLSX.
    """
    base_dir = Path(__file__).resolve().parent
    xlsx_path = base_dir / relative_dir / filename

    if not xlsx_path.exists():
        print(f"Файл не найден, пропускаю: {xlsx_path}")
        return

    wb = load_workbook(xlsx_path, data_only=False)

    result = {
        "file": str(xlsx_path),
        "sheets": [],
    }

    for ws in wb.worksheets:
        sheet_info = {
            "title": ws.title,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "rows": [],
            "formulas": [],
        }

        # Собираем содержимое строк (только непустые строки)
        for row in ws.iter_rows():
            cells = []
            has_value = False
            for cell in row:
                value = cell_repr(cell)
                cells.append({"coord": cell.coordinate, "value": value})
                if value not in (None, ""):
                    has_value = True
            if has_value:
                sheet_info["rows"].append(cells)

        # Собираем формулы
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    sheet_info["formulas"].append(
                        {
                            "coord": cell.coordinate,
                            "formula": v,
                        }
                    )

        result["sheets"].append(sheet_info)

    # Пишем результат в JSON, чтобы удобно смотреть и повторно использовать
    out_path = xlsx_path.with_suffix(".parsed.json")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"Готово. Структура и формулы сохранены в: {out_path}")


def main():
    # Кошка
    parse_xlsx("кошка", "ЭхоКГ протокол стандарт_КОШКА.xlsx")
    # Собака (будет пропущен, если файла пока нет)
    parse_xlsx("собака", "ЭхоКГ протокол стандарт_СОБАКА.xlsx")


if __name__ == "__main__":
    main()

